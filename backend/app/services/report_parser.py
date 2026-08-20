"""Report parser service: extract structured data from PDF, Excel, and CSV files.

Handles parsing of report attachments into structured DataTable format,
with timeout enforcement, size limits, and error handling for corrupted,
encrypted, or empty files.
"""

import csv
import io
import signal
import time
from pathlib import Path
from typing import Optional

import structlog

from app.models.schemas import DataTable, ParseResult, MAX_PARSE_SIZE_MB

logger = structlog.get_logger(__name__)

# --- Constants ---

PARSE_TIMEOUT_SECONDS = 120
MAX_PARSE_SIZE_BYTES = MAX_PARSE_SIZE_MB * 1024 * 1024


class ParseTimeoutError(Exception):
    """Raised when parsing exceeds the time limit."""
    pass


class ParseError(Exception):
    """Raised when a file cannot be parsed."""
    pass


# --- Timeout handling (Unix-style, graceful fallback on Windows) ---


class _TimeoutContext:
    """Context manager for enforcing parse timeout."""

    def __init__(self, seconds: int):
        self.seconds = seconds
        self._start_time: float = 0

    def __enter__(self):
        self._start_time = time.time()
        # signal.alarm only works on Unix; on Windows we rely on time checks
        try:
            signal.signal(signal.SIGALRM, self._handler)
            signal.alarm(self.seconds)
        except (AttributeError, ValueError):
            pass  # Windows or non-main thread
        return self

    def __exit__(self, *args):
        try:
            signal.alarm(0)
        except (AttributeError, ValueError):
            pass

    def _handler(self, signum, frame):
        raise ParseTimeoutError(
            f"Parsing exceeded {self.seconds} second timeout."
        )

    def check_timeout(self):
        """Manual timeout check for platforms without SIGALRM."""
        elapsed = time.time() - self._start_time
        if elapsed > self.seconds:
            raise ParseTimeoutError(
                f"Parsing exceeded {self.seconds} second timeout."
            )


# --- PDF Parser ---


def parse_pdf(file_path: str) -> ParseResult:
    """Parse a PDF file and extract tabular data using pdfplumber.

    Handles:
    - Corrupted files
    - Encrypted/password-protected files
    - Files with no data rows
    - 120-second timeout
    - 50 MB size limit

    Args:
        file_path: Path to the PDF file on disk.

    Returns:
        ParseResult with extracted tables or error info.
    """
    log = logger.bind(operation="parse_pdf", file_path=file_path)
    start_time = time.time()

    # Check file size (Req 2.6)
    path = Path(file_path)
    if not path.exists():
        return ParseResult(
            tables=[], file_type="pdf", parse_duration_seconds=0,
            success=False, error=f"File not found: {file_path}",
        )

    file_size = path.stat().st_size
    if file_size > MAX_PARSE_SIZE_BYTES:
        return ParseResult(
            tables=[], file_type="pdf", parse_duration_seconds=0,
            success=False,
            error=f"File size ({file_size / (1024*1024):.1f} MB) exceeds the {MAX_PARSE_SIZE_MB} MB limit.",
        )

    try:
        import pdfplumber

        with _TimeoutContext(PARSE_TIMEOUT_SECONDS) as timeout:
            try:
                pdf = pdfplumber.open(file_path)
            except Exception as e:
                error_msg = str(e).lower()
                if "password" in error_msg or "encrypt" in error_msg:
                    return ParseResult(
                        tables=[], file_type="pdf",
                        parse_duration_seconds=time.time() - start_time,
                        success=False,
                        error="File is password-protected or encrypted and cannot be accessed.",
                    )
                raise ParseError(f"Cannot open PDF: {e}")

            tables: list[DataTable] = []

            for page_num, page in enumerate(pdf.pages):
                timeout.check_timeout()
                extracted_tables = page.extract_tables()

                if not extracted_tables:
                    continue

                for table_idx, raw_table in enumerate(extracted_tables):
                    if not raw_table or len(raw_table) < 1:
                        continue

                    # First row as headers
                    headers = [str(h) if h else f"col_{i}" for i, h in enumerate(raw_table[0])]
                    rows = []
                    for row in raw_table[1:]:
                        rows.append([cell if cell is not None else "" for cell in row])

                    if rows:
                        tables.append(DataTable(
                            sheet_name=f"page_{page_num + 1}_table_{table_idx + 1}",
                            headers=headers,
                            rows=rows,
                            row_count=len(rows),
                            column_count=len(headers),
                        ))

            pdf.close()

        duration = time.time() - start_time

        # Check if no data found (Req 2.8)
        if not tables:
            log.info("pdf_no_data_rows", file_path=file_path)
            return ParseResult(
                tables=[], file_type="pdf",
                parse_duration_seconds=duration,
                success=False,
                error="No extractable data rows found in the file.",
            )

        log.info("pdf_parsed_successfully", tables_found=len(tables), duration_s=round(duration, 2))
        return ParseResult(
            tables=tables, file_type="pdf",
            parse_duration_seconds=duration,
            success=True,
        )

    except ParseTimeoutError as e:
        return ParseResult(
            tables=[], file_type="pdf",
            parse_duration_seconds=PARSE_TIMEOUT_SECONDS,
            success=False, error=str(e),
        )
    except ParseError as e:
        return ParseResult(
            tables=[], file_type="pdf",
            parse_duration_seconds=time.time() - start_time,
            success=False, error=str(e),
        )
    except Exception as e:
        log.error("pdf_parse_error", error=str(e), exc_info=True)
        return ParseResult(
            tables=[], file_type="pdf",
            parse_duration_seconds=time.time() - start_time,
            success=False, error=f"Failed to parse PDF: {e}",
        )


# --- Excel Parser ---


def parse_excel(file_path: str) -> ParseResult:
    """Parse an Excel file (.xlsx/.xls) and extract cell values from all sheets.

    Handles:
    - Corrupted files
    - Password-protected workbooks
    - Preserves sheet names and row/column structure
    - Extracts resolved cell values (not formulas)

    Args:
        file_path: Path to the Excel file on disk.

    Returns:
        ParseResult with extracted tables or error info.
    """
    log = logger.bind(operation="parse_excel", file_path=file_path)
    start_time = time.time()

    path = Path(file_path)
    if not path.exists():
        return ParseResult(
            tables=[], file_type="excel", parse_duration_seconds=0,
            success=False, error=f"File not found: {file_path}",
        )

    file_size = path.stat().st_size
    if file_size > MAX_PARSE_SIZE_BYTES:
        return ParseResult(
            tables=[], file_type="excel", parse_duration_seconds=0,
            success=False,
            error=f"File size ({file_size / (1024*1024):.1f} MB) exceeds the {MAX_PARSE_SIZE_MB} MB limit.",
        )

    try:
        from openpyxl import load_workbook
        from openpyxl.utils.exceptions import InvalidFileException

        try:
            # data_only=True to get resolved values instead of formulas (Req 2.2)
            wb = load_workbook(file_path, data_only=True, read_only=True)
        except InvalidFileException as e:
            return ParseResult(
                tables=[], file_type="excel",
                parse_duration_seconds=time.time() - start_time,
                success=False, error=f"Corrupted or invalid Excel file: {e}",
            )
        except Exception as e:
            error_msg = str(e).lower()
            if "password" in error_msg or "encrypt" in error_msg or "protected" in error_msg:
                return ParseResult(
                    tables=[], file_type="excel",
                    parse_duration_seconds=time.time() - start_time,
                    success=False,
                    error="File is password-protected or encrypted and cannot be accessed.",
                )
            raise

        tables: list[DataTable] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_data = []

            for row in ws.iter_rows(values_only=True):
                # Convert all cell values to strings (preserving None as empty)
                row_values = [str(cell) if cell is not None else "" for cell in row]
                # Skip completely empty rows
                if any(v.strip() for v in row_values):
                    rows_data.append(row_values)

            if not rows_data:
                continue

            # First non-empty row as headers
            headers = rows_data[0]
            data_rows = rows_data[1:]

            if data_rows:
                tables.append(DataTable(
                    sheet_name=sheet_name,
                    headers=headers,
                    rows=data_rows,
                    row_count=len(data_rows),
                    column_count=len(headers),
                ))

        wb.close()
        duration = time.time() - start_time

        if not tables:
            log.info("excel_no_data_rows", file_path=file_path)
            return ParseResult(
                tables=[], file_type="excel",
                parse_duration_seconds=duration,
                success=False,
                error="No extractable data rows found in the file.",
            )

        log.info("excel_parsed_successfully", sheets=len(tables), duration_s=round(duration, 2))
        return ParseResult(
            tables=tables, file_type="excel",
            parse_duration_seconds=duration,
            success=True,
        )

    except Exception as e:
        log.error("excel_parse_error", error=str(e), exc_info=True)
        return ParseResult(
            tables=[], file_type="excel",
            parse_duration_seconds=time.time() - start_time,
            success=False, error=f"Failed to parse Excel file: {e}",
        )


# --- CSV Parser ---


def detect_delimiter(sample: str) -> str:
    """Auto-detect CSV delimiter from a file sample.

    Supports comma, semicolon, and tab delimiters (Req 2.3).
    Uses Python's csv.Sniffer, falling back to frequency analysis.

    Args:
        sample: A string sample from the beginning of the CSV file.

    Returns:
        The detected delimiter character.
    """
    # Try csv.Sniffer first
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        return dialect.delimiter
    except csv.Error:
        pass

    # Fallback: frequency analysis
    counts = {
        ",": sample.count(","),
        ";": sample.count(";"),
        "\t": sample.count("\t"),
    }

    # Return the delimiter with highest count; default to comma
    best = max(counts, key=counts.get)
    if counts[best] == 0:
        return ","
    return best


def parse_csv(file_path: str) -> ParseResult:
    """Parse a CSV file with auto-delimiter detection.

    Supports comma, semicolon, and tab delimiters.
    Handles empty files and malformed CSV.

    Args:
        file_path: Path to the CSV file on disk.

    Returns:
        ParseResult with extracted table or error info.
    """
    log = logger.bind(operation="parse_csv", file_path=file_path)
    start_time = time.time()

    path = Path(file_path)
    if not path.exists():
        return ParseResult(
            tables=[], file_type="csv", parse_duration_seconds=0,
            success=False, error=f"File not found: {file_path}",
        )

    file_size = path.stat().st_size
    if file_size > MAX_PARSE_SIZE_BYTES:
        return ParseResult(
            tables=[], file_type="csv", parse_duration_seconds=0,
            success=False,
            error=f"File size ({file_size / (1024*1024):.1f} MB) exceeds the {MAX_PARSE_SIZE_MB} MB limit.",
        )

    try:
        # Read file content
        content = path.read_text(encoding="utf-8", errors="replace")

        if not content.strip():
            return ParseResult(
                tables=[], file_type="csv",
                parse_duration_seconds=time.time() - start_time,
                success=False,
                error="No extractable data rows found in the file.",
            )

        # Auto-detect delimiter using sample (first 8KB)
        sample = content[:8192]
        delimiter = detect_delimiter(sample)
        log.info("csv_delimiter_detected", delimiter=repr(delimiter))

        # Parse CSV
        reader = csv.reader(io.StringIO(content), delimiter=delimiter)
        all_rows = []
        for row in reader:
            # Skip completely empty rows
            if any(cell.strip() for cell in row):
                all_rows.append(row)

        if not all_rows:
            return ParseResult(
                tables=[], file_type="csv",
                parse_duration_seconds=time.time() - start_time,
                success=False,
                error="No extractable data rows found in the file.",
            )

        # First row as headers
        headers = all_rows[0]
        data_rows = all_rows[1:]

        if not data_rows:
            return ParseResult(
                tables=[], file_type="csv",
                parse_duration_seconds=time.time() - start_time,
                success=False,
                error="No extractable data rows found in the file.",
            )

        # Normalize row lengths to match header count
        col_count = len(headers)
        normalized_rows = []
        for row in data_rows:
            if len(row) < col_count:
                row = row + [""] * (col_count - len(row))
            elif len(row) > col_count:
                row = row[:col_count]
            normalized_rows.append(row)

        table = DataTable(
            sheet_name=None,
            headers=headers,
            rows=normalized_rows,
            row_count=len(normalized_rows),
            column_count=col_count,
        )

        duration = time.time() - start_time
        log.info("csv_parsed_successfully", rows=len(normalized_rows), duration_s=round(duration, 2))

        return ParseResult(
            tables=[table], file_type="csv",
            parse_duration_seconds=duration,
            success=True,
        )

    except Exception as e:
        log.error("csv_parse_error", error=str(e), exc_info=True)
        return ParseResult(
            tables=[], file_type="csv",
            parse_duration_seconds=time.time() - start_time,
            success=False, error=f"Failed to parse CSV file: {e}",
        )


# --- Router: dispatch to correct parser ---


def parse_report(file_path: str, file_type: str) -> ParseResult:
    """Route to the correct parser based on file type.

    Args:
        file_path: Path to the file on disk.
        file_type: File extension without dot (pdf, xlsx, xls, csv).

    Returns:
        ParseResult from the appropriate parser.
    """
    file_type = file_type.lower().strip(".")

    if file_type == "pdf":
        return parse_pdf(file_path)
    elif file_type in ("xlsx", "xls"):
        return parse_excel(file_path)
    elif file_type == "csv":
        return parse_csv(file_path)
    else:
        return ParseResult(
            tables=[], file_type=file_type, parse_duration_seconds=0,
            success=False, error=f"Unsupported file type: {file_type}",
        )
